import requests
import urllib.request
import openpyxl
import os





class Product:
    brokerId= 0
    available= True
    sold= False
    qualification= 0
    propertyLink= ''
    saleRent= 0
    propertyType= 0
    deliveryType= 0
    latitude= 0
    longitude= 0
    zone= 0
    location= ''
    adress= ''
    street= 0
    career= 0
    administrationValue= 0
    saleValue= 0
    roomQuantity= 0
    floorQuantity= 0
    bathroomQuantity= 0
    parkingQuantity= 0
    independentParkingQuantity= 0
    serviceParkingQuantity= 0
    builtArea= 0
    privateArea= 0
    terraceBalconyArea= 0
    year= 0
    stratus= 0
    view= 0
    floor= 0
    commission= 0
    description= ''
    comments= ''
    commonZones= ''

	

def writeProperty(property):
	print('{')
	print('		"qradorId":' + str(2) + ',')
	print('		"brokerId":' + str(2) + ',')
	print('		"available":' + str('boolean') + ',')
	print('		"sold":' + str('boolean') + ',')
	print('		"qualification":' + str(2) + ',')
	print('		"propertyLink":' + 'string' + ',')
	print('		"saleRent":' + str(2) + ',')
	print('		"propertyType":' + str(2) + ',')
	print('		"deliveryType":' + str(2) + ',')
	print('		"latitude":' + str(2) + ',')
	print('		"longitude":' + str(2) + ',')
	print('		"zone":' + str(2) + ',')
	print('		"location":' + 'string' +',')
	print('		"adress":' + 'string' +',')
	print('		"street":' + str(2) + ',')
	print('		"career":' + str(2) + ',')
	print('		"administrationValue":' + str(2) + ',')
	print('		"saleValue":' + str(2) + ',')
	print('		"roomQuantity":' + str(2) + ',')
	print('		"floorQuantity":' + str(2) + ',')
	print('		"bathroomQuantity":' +str(2) + ',')
	print('		"parkingQuantity":' + str(2) + ',')
	print('		"independentParkingQuantity":' + str(2) + ',')
	print('		"serviceParkingQuantity":' + str(2) + ',')
	print('		"builtArea":' + str(2) + ',')
	print('		"privateArea":' + str(2) + ',')
	print('		"terraceBalconyArea":' + str(2) + ',')
	print('		"year": '+ str(2) + ',')
	print('		"stratus":' + str(2) + ',')
	print('		"view":' + str(2) + ',')
	print('		"floor":' + str(2) + ',')
	print('		"commission":' + str(2) + ',')
	print('		"description":' + 'string' +',')
	print('		"comments":' + 'string' +',')
	print('		"commonZones":' + 'string' +',')
	print('}')


def readProperties():
	adressFormats = {
        1: 'Cl. 93 # 9-46',
        2: 'Cra 5 # 78 - 20',
        3: 'Calle 92 # 14 - 23 Apto 1002',
        4: 'Cr. 18a # 106-16',
        5: 'Carrera 19C #85-64',
		6: '94 A con 21',
		7: 'Calle 74 con 2',
		8: 'Cll 131 con 19',
		9: 'Cra 23 con 104',
		10: 'cra4 con 76',
		11: 'Tr 23 con 93',

    }


def readProperties():
	qualificator = {
        "Muy bueno": '"Muy bueno"',
        "Bueno": '"Bueno"',
        "Regular": '"Regular"',
        "Bueno para remodelar": '"Bueno para remodelar"',
        "Malo": '"Malo"',
		"Remodelar": '"Remodelar"',

    }
	saleRenter = {
		"venta":  '"Venta"',
		"arriendo": '"Arriendo"',
		"arriendo/venta": '"Venta o Arriendo"',
    }  
	propertyTyper = {
		"Apartamento": '"Apartamento"',
		"apartamento": '"Apartamento"',
		"Apartaestudio": '"Apartamento"',
		"Proyecto": '"Apartamento"',
		"Casa": '"Casa"',
		"casa": '"Casa"',
    }   
	deliveryTyper = {
		"Usado": '"Usado"',
		"Inmediata": '"Usado"',
		"Nuevo entrega inmediata": '"Nuevo con entrega inmediata"',
		"Sobre Planos": '"Sobre planos"',		
    }
	zoner = {
		"#N/A": 'null',
		"#N/D": 'null',
		"Chicó": '"Chicó"',
		"Multicentro": '"Multicentro"',
		"Chapinero": '"Chapinero"',
		"Córdoba": '"Córdoba"',		
		"Teusaquillo": '"Teusaquillo"',		
		"La Carolina": '"La Carolina"',
		"Chapinero Alto": '"Chapinero Alto"',
		"Mazurén": '"Mazurén"',
		"Tunal": '"Tunal"',		
		"Cedritos": '"Cedritos"',
		"La Castellana": '"La Castellana"',
		"Suba": '"Suba"',	

    }  
	viewer = {
		"Interior": '"Interior"',
		"Conjunto": '"Interior"',
		"Lateral": '"Interior"',
		"Exterior": '"Exterior"',
		"Exterior e Interior": '"Exterior"',
		"Esquinero Interior": '"Esquinero Interior"',
		"Esquinero Exterior": '"Esquinero Exterior"',
    }

	path = 'C:\\Users\\jdbar\\Documents\\Qrador\\Propiedades3.xlsx'


	wb = openpyxl.load_workbook(path, data_only=True)
	sheet = wb.active 
	
	print('[')

	i = 2
	for x in range (100):
		qradorId = 'null' if sheet.cell(row = (x+i), column = 1).value is None else sheet.cell(row = (x+i), column = 1).value[3:]
		available = 'true' if sheet.cell(row = (x+i), column = 3).value is None else 'true' if sheet.cell(row = (x+i), column = 3).value == 'Sí' else 'false'
		sold = 'false' if available == 'true'else 'true'
		qualification = 'null' if sheet.cell(row = (x+i), column = 5).value is None else qualificator.get(sheet.cell(row = (x+i), column = 5).value, 'null')
		brokerId = 'null' if sheet.cell(row = (x+i), column = 9).value is None else sheet.cell(row = (x+i), column = 9).value
		propertyLink = 'null' if sheet.cell(row = (x+i), column = 10).value is None else '"'+sheet.cell(row = (x+i), column = 10).value+'"'
		view360 = 'null' if sheet.cell(row = (x+i), column = 11).value is None else '"'+sheet.cell(row = (x+i), column = 11).value+'"'
		saleRent = '"Venta"' if sheet.cell(row = (x+i), column = 12).value is None else saleRenter.get(sheet.cell(row = (x+i), column = 12).value, '"Venta"')
		propertyType = '"Apartamento"' if sheet.cell(row = (x+i), column = 13).value is None else propertyTyper.get(sheet.cell(row = (x+i), column = 13).value, '"Apartamento"')
		deliveryType = '"Usado"' if sheet.cell(row = (x+i), column = 14).value is None else deliveryTyper.get(sheet.cell(row = (x+i), column = 14).value, '"Usado"')
		zone = 'null' if sheet.cell(row = (x+i), column = 15).value is None else zoner.get(sheet.cell(row = (x+i), column = 15).value, 'null') 
		location = 'null' if sheet.cell(row = (x+i), column = 16).value is None else '"'+sheet.cell(row = (x+i), column = 16).value+'"' 
		latitude = 'null' if sheet.cell(row = (x+i), column = 17).value is None or sheet.cell(row = (x+i), column = 17).value == '#N/A' or sheet.cell(row = (x+i), column = 17).value == 0 else sheet.cell(row = (x+i), column = 17).value.split(',')[0]
		longitude = 'null' if sheet.cell(row = (x+i), column = 17).value is None or sheet.cell(row = (x+i), column = 17).value == '#N/A' or sheet.cell(row = (x+i), column = 17).value == 0 else sheet.cell(row = (x+i), column = 17).value.split(',')[1]  
		adress = 'null' if sheet.cell(row = (x+i), column = 18).value is None else '"'+sheet.cell(row = (x+i), column = 18).value+'"'
		street = 'null' if sheet.cell(row = (x+i), column = 19).value is None else sheet.cell(row = (x+i), column = 19).value
		career = 'null' if sheet.cell(row = (x+i), column = 20).value is None else sheet.cell(row = (x+i), column = 20).value 
		leaseValue = 'null' if sheet.cell(row = (x+i), column = 21).value is None else sheet.cell(row = (x+i), column = 21).value
		leaseAdminValue = 'null' if sheet.cell(row = (x+i), column = 22).value is None else sheet.cell(row = (x+i), column = 22).value
		administrationValue = 'null' if sheet.cell(row = (x+i), column = 23).value is None else sheet.cell(row = (x+i), column = 23).value 
		saleValue = 'null' if sheet.cell(row = (x+i), column = 24).value is None else sheet.cell(row = (x+i), column = 24).value 
		roomQuantity = 'null' if sheet.cell(row = (x+i), column = 25).value is None else sheet.cell(row = (x+i), column = 25).value
		bathroomQuantity = 'null' if sheet.cell(row = (x+i), column = 26).value is None else sheet.cell(row = (x+i), column = 26).value
		parkingQuantity = 'null' if sheet.cell(row = (x+i), column = 27).value is None else sheet.cell(row = (x+i), column = 27).value
		parkingType = 'null' if sheet.cell(row = (x+i), column = 28).value is None else '"'+sheet.cell(row = (x+i), column = 28).value+'"'
		builtArea = 'null' if sheet.cell(row = (x+i), column = 29).value is None else sheet.cell(row = (x+i), column = 29).value
		privateArea = 'null' if sheet.cell(row = (x+i), column = 30).value is None else sheet.cell(row = (x+i), column = 30).value
		terraceBalconyDescription = 'null' if sheet.cell(row = (x+i), column = 31).value is None else '"'+sheet.cell(row = (x+i), column = 31).value+'"'
		terraceBalconyArea = 'null' if sheet.cell(row = (x+i), column = 32).value is None else sheet.cell(row = (x+i), column = 32).value
		terraceArea  = 'null' if sheet.cell(row = (x+i), column = 33).value is None else '['+str(sheet.cell(row = (x+i), column = 33).value)+']'
		balconyArea = 'null' if sheet.cell(row = (x+i), column = 34).value is None else '['+str(sheet.cell(row = (x+i), column = 34).value)+']'
		backyardArea = 'null' if sheet.cell(row = (x+i), column = 35).value is None else '['+str(sheet.cell(row = (x+i), column = 35).value)+']'
		year = 'null' if sheet.cell(row = (x+i), column = 36).value is None else sheet.cell(row = (x+i), column = 36).value
		stratus = 'null' if sheet.cell(row = (x+i), column = 37).value is None else sheet.cell(row = (x+i), column = 37).value
		view = 'null' if sheet.cell(row = (x+i), column = 38).value is None else viewer.get(sheet.cell(row = (x+i), column = 38).value, 'null')
		floor = 'null' if sheet.cell(row = (x+i), column = 39).value is None else sheet.cell(row = (x+i), column = 39).value
		commission = 'null' if sheet.cell(row = (x+i), column = 40).value is None else sheet.cell(row = (x+i), column = 40).value
		description = 'null' if sheet.cell(row = (x+i), column = 41).value is None else '"'+sheet.cell(row = (x+i), column = 41).value+'"'
		comments = 'null' if sheet.cell(row = (x+i), column = 42).value is None else '"'+sheet.cell(row = (x+i), column = 42).value+'"'
		print('{')
		print('  "id":' + str(qradorId) + ',')
		print('  "brokerId":' + str(brokerId) + ',')
		print('  "available":' + available.replace('\n', ' ').replace('\r', ' ') + ',')
		print('  "sold":' + sold.replace('\n', ' ').replace('\r', ' ') + ',')
		
		print('  "dealType":' + saleRent + ',')
		print('  "propertyType":' + propertyType + ',')
		print('  "deliveryType":' + deliveryType + ',')
		print('  "rating":' + qualification + ',')

		print('  "latitude":' + str(latitude) + ',')
		print('  "longitude":' + str(longitude) + ',')
		print('  "zone":' + zone + ',')
		print('  "location":' + location.replace('\n', ' ').replace('\r', ' ') +',')
		print('  "address":' + adress.replace('\n', ' ').replace('\r', ' ') +',')
		print('  "street":' + str(street) + ',')
		print('  "avenue":' + str(career) + ',')

		print('  "price":' + str(saleValue) + ',')
		print('  "monthlyMaintenanceFee":' + str(administrationValue) + ',')
		print('  "rentPrice":' + str(leaseValue) + ',')
		print('  "rentWithMonthlyMaintenancePrice":' + str(leaseAdminValue) + ',')
		print('  "commissionPercent":' + str(commission) + ',')

		print('  "carpetArea":' + str(privateArea) + ',')
		print('  "builtUpArea":' + str(builtArea) + ',')
		print('  "terraceBalconyDescription":' + terraceBalconyDescription.replace('\n', ' ').replace('\r', ' ') +',')
		print('  "terraceBalconyArea":' + str(terraceBalconyArea) + ',')
		print('  "terraceArea":' + str(terraceArea) + ',')
		print('  "balconyArea":' + str(balconyArea) + ',')
		print('  "backyardArea":' + str(backyardArea) + ',')

		print('  "floor":' + str(floor) + ',')
		print('  "propertyFloorCount": null,')
		print('  "rooms":' + str(roomQuantity) + ',')
		print('  "bathrooms":' +str(bathroomQuantity) + ',')
		print('  "parkingSpaces":' + str(parkingQuantity) + ',')
		print('  "parkingType":' + parkingType.replace('\n', ' ').replace('\r', ' ') + ',')
		print('  "independentParkingSpaces": null,')
		print('  "serviceParkingSpaces": null,')
		print('  "view":' + view + ',')

		print('  "constructionYear": '+ str(year) + ',')
		print('  "commonAreas": null,')
		print('  "estrato":' + str(stratus) + ',')

		print('  "description":' + description.replace('\n', ' ').replace('\r', ' ') +',')
		print('  "comments":' + comments.replace('\n', ' ').replace('\r', ' ') +',')
		print('  "sourceLink":' + propertyLink.replace('\n', ' ').replace('\r', ' ') + ',')
		print('  "tour360Link":' + view360.replace('\n', ' ').replace('\r', ' '))
		
		print('},')

	print(']')




readProperties()





